import openpyxl #if not installed pip install openpyxl
import os

#this program illustrates reading of excel sheet using python
#some of the features may be depriated now so need to change accordingly
workbook = openpyxl.load_workbook('example.xlsx') #it is used to declare Excel file
#type(workbook)
#sheet = workbook.get_sheet_by_name('Sheet1') #it will designate our current Excel file  This function is depricated use one in below
sheet = workbook['Sheet1']  #new method for getting sheets
#type(sheet)
#print(workbook.get_sheet_names()) Depricated feature
print(workbook.sheetnames)

cell = sheet['A1']
print(cell.value) #will print value
print(sheet['B3'].value)

cell=sheet.cell(row=1 , column=2)
print(cell)
print(cell.value)