import openpyxl,os
#opening and editing excel sheet or creating excel sheets

wb = openpyxl.Workbook() # will create the the new workbook in memory
print(wb.sheetnames) #will print sheet name in this new workbook

sheet = wb['Sheet']  #sheet variable  

sheet['A1']= 'Roll Number' #adding the variable using assignment statements
sheet['A2'] = 145678
sheet['B1'] = 'Name'
sheet['B2'] = 'C k nayar'

print(sheet['B1'].value) #printing the value to check

wb.save('example2.xlsx') # will save in hard Drive here i have given releavtive path

sheet2=wb.create_sheet('Sheet1') #even if argument 'Sheet1' not passed still it can work perfectly
print(wb.sheetnames)
print(sheet2.title)
#even we can change sheet name
sheet2.title = 'new_sheet1'
print(wb.sheetnames) # to see changes
#wb.save('example2.xlsx')

sheet3=wb.create_sheet('my_indexed_sheet',index=0) #this sheet will come at very begining
#wb.remove_sheet('')
print(wb.sheetnames) # to see changes
wb.save('example2.xlsx')






